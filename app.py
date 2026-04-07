"""
AGE_24 Web Processor
====================
Run this once on the office PC:
    pip install flask pandas openpyxl
    python app.py

Then anyone on the same WiFi can open:
    http://<this-pc-ip>:5000
"""

import sys
import shutil
import datetime
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from flask import Flask, request, send_file, render_template, jsonify

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB max upload

# ── Constants ─────────────────────────────────────────────────────────────────
EXCLUDE_ALL  = {"11-DMERC", "1-Self Pay", "20-Work Comp", "23-Auto", "9-EPS"}
WC_MVA_CATS  = {"20-Work Comp", "23-Auto"}
AMOUNT_COLS  = ["CurrentAmt", "Over30Amt", "Over60Amt", "Over90Amt", "Over120Amt", "Over150Amt"]
PCT_ROWS     = {10, 18, 26, 35, 43, 51}
SUMMARY_SHEET = "Summary"

# ── Helpers ───────────────────────────────────────────────────────────────────
def bucket_sums(df, mask):
    sub = df.loc[mask, AMOUNT_COLS].apply(pd.to_numeric, errors="coerce").fillna(0)
    return [float(sub[c].sum()) for c in AMOUNT_COLS]

def total_ar(vals):
    return sum(vals)

def pct_over_120(vals):
    t = total_ar(vals)
    return (vals[4] + vals[5]) / t if t else 0.0

def find_date_column(ws, target_date):
    anchor_val = ws.cell(row=1, column=3).value
    if anchor_val is None:
        return None
    if hasattr(anchor_val, "date"):
        anchor_date = anchor_val.date()
    elif isinstance(anchor_val, datetime.datetime):
        anchor_date = anchor_val.date()
    else:
        try:
            anchor_date = datetime.datetime.strptime(str(anchor_val), "%Y-%m-%d").date()
        except ValueError:
            return None
    delta = (target_date - anchor_date).days
    if delta < 0:
        return None
    return 3 + delta

def detect_raw_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == SUMMARY_SHEET.lower():
            continue
        try:
            datetime.datetime.strptime(name.strip(), "%m-%d-%Y")
            return name
        except ValueError:
            continue
    return None

# ── Core processor ────────────────────────────────────────────────────────────
def process_bytes(file_bytes, start_date=None, end_date=None):
    """
    Process Excel bytes in memory, return:
        (success, message, output_bytes_or_none, summary_dict)

    Parameters
    ----------
    file_bytes  : raw bytes of the uploaded .xlsx file
    start_date  : optional str in 'YYYY-MM-DD' format (from HTML date picker)
    end_date    : optional str in 'YYYY-MM-DD' format (from HTML date picker)
    """
    buf = io.BytesIO(file_bytes)

    try:
        wb_check = load_workbook(buf, read_only=True, data_only=True)
        raw_sheet_name = detect_raw_sheet(wb_check)
        wb_check.close()
    except Exception as e:
        return False, f"Cannot open file: {e}", None, None

    if raw_sheet_name is None:
        return False, (
            "No raw data sheet found. Sheet name must be a date in MM-DD-YYYY format "
            "(e.g. 04-06-2026)."
        ), None, None

    sheet_date = datetime.datetime.strptime(raw_sheet_name.strip(), "%m-%d-%Y").date()

    try:
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name=raw_sheet_name, header=0)
    except Exception as e:
        return False, f"Cannot read sheet '{raw_sheet_name}': {e}", None, None

    if "textbox18" not in df.columns:
        return False, "Column 'textbox18' not found in raw data sheet.", None, None

    missing = [c for c in AMOUNT_COLS + ["Financial_Class"] if c not in df.columns]
    if missing:
        return False, f"Missing columns in raw data: {missing}", None, None

    # ── Filter by positive textbox18 ──────────────────────────────────────────
    df = df[pd.to_numeric(df["textbox18"], errors="coerce") > 0].copy()

    # ── Date range filter on svc_date ─────────────────────────────────────────
    date_filter_applied = False
    date_filter_label   = None

    if start_date or end_date:
        if "svc_date" not in df.columns:
            return False, (
                "Column 'svc_date' not found in the raw data sheet. "
                "Date range filtering requires this column."
            ), None, None

        # Normalise svc_date – handles mm/dd/yyyy, yyyy-mm-dd, Excel serial dates, etc.
        df["svc_date"] = pd.to_datetime(df["svc_date"], errors="coerce")

        before = len(df)
        if start_date:
            start_dt = pd.to_datetime(start_date)   # 'YYYY-MM-DD' from HTML picker
            df = df[df["svc_date"] >= start_dt]
        if end_date:
            end_dt = pd.to_datetime(end_date)        # 'YYYY-MM-DD' from HTML picker
            df = df[df["svc_date"] <= end_dt]

        after = len(df)
        date_filter_applied = True
        date_filter_label   = f"{start_date or '—'}  →  {end_date or '—'}  ({after} of {before} rows)"

    filtered_count = len(df)
    fc = df["Financial_Class"]

    masks = {
        "Insurance": ~fc.isin(EXCLUDE_ALL),
        "Patient":    fc == "1-Self Pay",
        "WC":         fc.isin(WC_MVA_CATS),
        "EPS":        fc == "9-EPS",
        "DME":        fc == "11-DMERC",
    }
    res = {k: bucket_sums(df, v) for k, v in masks.items()}
    overall = [sum(res[k][i] for k in res) for i in range(6)]

    buf.seek(0)
    wb = load_workbook(buf)

    if SUMMARY_SHEET not in wb.sheetnames:
        return False, f"Sheet '{SUMMARY_SHEET}' not found in workbook.", None, None

    ws = wb[SUMMARY_SHEET]
    date_col = find_date_column(ws, sheet_date)
    if date_col is None:
        return False, (
            f"Date {sheet_date} not found in Summary sheet header row. "
            "Make sure the Summary sheet covers this date."
        ), None, None

    col_letter = get_column_letter(date_col)

    updates = {
        3: total_ar(overall),    4: overall[0],  5: overall[1],
        6: overall[2],           7: overall[3],  8: overall[4],
        9: overall[5],          10: pct_over_120(overall),
        11: total_ar(res["Insurance"]), 12: res["Insurance"][0], 13: res["Insurance"][1],
        14: res["Insurance"][2], 15: res["Insurance"][3], 16: res["Insurance"][4],
        17: res["Insurance"][5], 18: pct_over_120(res["Insurance"]),
        19: total_ar(res["Patient"]), 20: res["Patient"][0], 21: res["Patient"][1],
        22: res["Patient"][2],   23: res["Patient"][3], 24: res["Patient"][4],
        25: res["Patient"][5],   26: pct_over_120(res["Patient"]),
        27: total_ar(res["WC"]), 28: total_ar(res["WC"]), 29: res["WC"][0],
        30: res["WC"][1],        31: res["WC"][2], 32: res["WC"][3],
        33: res["WC"][4],        34: res["WC"][5], 35: pct_over_120(res["WC"]),
        36: total_ar(res["EPS"]),37: res["EPS"][0], 38: res["EPS"][1],
        39: res["EPS"][2],       40: res["EPS"][3], 41: res["EPS"][4],
        42: res["EPS"][5],       43: pct_over_120(res["EPS"]),
        44: total_ar(res["DME"]),45: res["DME"][0], 46: res["DME"][1],
        47: res["DME"][2],       48: res["DME"][3], 49: res["DME"][4],
        50: res["DME"][5],       51: pct_over_120(res["DME"]),
    }

    for row, val in updates.items():
        cell = ws.cell(row=row, column=date_col)
        cell.value = round(float(val), 4)
        cell.number_format = "0.00%" if row in PCT_ROWS else "#,##0.00"

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    summary = {
        "sheet": raw_sheet_name,
        "date": str(sheet_date),
        "col_letter": col_letter,
        "rows": filtered_count,
        "date_filter": date_filter_label,   # None when no filter was applied
        "categories": [
            {"name": "Overall Aging",  "total": round(total_ar(overall), 2),           "pct": round(pct_over_120(overall)*100, 2)},
            {"name": "Insurance AR",   "total": round(total_ar(res["Insurance"]), 2),  "pct": round(pct_over_120(res["Insurance"])*100, 2)},
            {"name": "Patient AR",     "total": round(total_ar(res["Patient"]), 2),    "pct": round(pct_over_120(res["Patient"])*100, 2)},
            {"name": "WC & MVA",       "total": round(total_ar(res["WC"]), 2),         "pct": round(pct_over_120(res["WC"])*100, 2)},
            {"name": "EPS AR",         "total": round(total_ar(res["EPS"]), 2),        "pct": round(pct_over_120(res["EPS"])*100, 2)},
            {"name": "DME AR",         "total": round(total_ar(res["DME"]), 2),        "pct": round(pct_over_120(res["DME"])*100, 2)},
        ]
    }

    return True, "Success", out_buf.read(), summary

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/process", methods=["POST"])
def process_route():
    if "file" not in request.files:
        return jsonify({"success": False, "message": "No file uploaded."}), 400

    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify({"success": False, "message": "Please upload an .xlsx file."}), 400

    # ── Read optional date range fields sent from the HTML form ───────────────
    start_date = request.form.get("start_date") or None   # 'YYYY-MM-DD' or empty string
    end_date   = request.form.get("end_date")   or None   # 'YYYY-MM-DD' or empty string

    file_bytes = f.read()
    success, message, out_bytes, summary = process_bytes(file_bytes, start_date, end_date)

    if not success:
        return jsonify({"success": False, "message": message}), 400

    # Store output in app context temporarily (keyed by session-ish token)
    import hashlib, time
    token = hashlib.md5(f"{time.time()}".encode()).hexdigest()[:12]
    app.config.setdefault("_outputs", {})[token] = (out_bytes, f.filename)

    return jsonify({"success": True, "token": token, "summary": summary})

@app.route("/download/<token>")
def download(token):
    outputs = app.config.get("_outputs", {})
    if token not in outputs:
        return "File not found or expired.", 404
    out_bytes, orig_name = outputs[token]
    stem = Path(orig_name).stem
    out_name = f"{stem}_updated.xlsx"
    return send_file(
        io.BytesIO(out_bytes),
        as_attachment=True,
        download_name=out_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    import socket
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except:
        local_ip = "your-pc-ip"
    print(f"\n{'='*55}")
    print(f"  AGE_24 Web Processor is running!")
    print(f"  Open in browser on this PC : http://localhost:5000")
    print(f"  Open from other PCs on WiFi: http://{local_ip}:5000")
    print(f"{'='*55}\n")
    app.run(host="0.0.0.0", port=5000, debug=False)
